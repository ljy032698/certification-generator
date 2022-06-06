import pandas as pd
import docx
from docx.oxml.ns import qn
from docx.enum.text import WD_ALIGN_PARAGRAPH
from openpyxl import load_workbook
from openpyxl import Workbook

def fetch_data_from_excel_file(sheet):
    name_list = []
    registration_num_list = []
    address_list = []
    division_list = []
    position_list = []
    employment_period_list = []
    dutiees_list = []

    for i in range(1, sheet.max_row + 1):
        name_list.append(sheet.cell(i, 1).value)
        registration_num_list.append(sheet.cell(i, 2).value)
        address_list.append(sheet.cell(i, 3).value)
        division_list.append(sheet.cell(i, 4).value)
        position_list.append(sheet.cell(i, 5).value)
        employment_period_list.append(sheet.cell(i, 6).value)
        dutiees_list.append(sheet.cell(i, 7).value)
    
    # print
    print(name_list)
    print(registration_num_list)
    print(address_list)
    print(division_list)
    print(position_list)
    print(employment_period_list)
    print(dutiees_list)

def generator_excel_file():
    df = pd.DataFrame([["성명", "주민등록번호", "주소", "소속", "직위", "재직기간", "담당업무"],
                    ["김민준", "900506-1581223", "서울특별시 강남구 역삼동 824-25", "R&D", "대리", "2020.10.14~2022.06.06", "소프트웨어엔지니어"],
                    ["김철수", "940808-1042932", "서울특별시 강남구 강남대로 92길 1", "HR", "사원", "2019.04.02~2022.06.06", "HR매니저"],
                    ["김영희", "900909-2039423", "서울특별시 관악구 관악로6길 2", "R&D", "부장", "2017.03.28~2022.06.06", "하드웨어매니저"],
                    ["이서준", "921010-1204281", "서울특별시 동작구 국사봉2길 1", "AD", "대리", "2018.08.18~2022.06.06", "회계매니저"],
                    ["장다인", "001212-4028324", "서울특별시 구로구 도림로100", "AD", "사원", "2022.01.05~2022.06.06", "회계매니저" ],])
    df.to_excel("personal_info.xlsx", index=False, header=False)

try:
    wb = load_workbook("personal_info.xlsx", data_only=True)
    sheet = wb.active
except:
    wb = Workbook()
    sheet = wb.active
    generator_excel_file()

fetch_data_from_excel_file(sheet)

wb.save("personal_info.xlsx")